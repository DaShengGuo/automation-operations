"""
comment_bot/materials.py
Excel 素材管理器 — 文案/图片/私信/回复 的读取和随机选取
支持模板随机变体生成，避免内容重复
"""
from __future__ import annotations

import random
import re
import time
from pathlib import Path
from typing import Optional

import openpyxl

from douyin_core import config as cfg


# ── 模板变体引擎 ──

# 同义词替换表：用于随机生成模板变体
SYNONYM_MAP = {
    "吃药": ["吃药", "口服药", "用药", "服药"],
    "照光": ["照光", "光照", "光疗", "光治疗"],
    "光照": ["光照", "照光", "光疗", "光治疗"],
    "弄好": ["弄好", "治好", "恢复", "康复"],
    "恢复": ["恢复", "康复", "好转", "变好"],
    "坚持": ["坚持", "持续", "一直", "不懈"],
    "真的": ["真的", "确实", "真心", "实在"],
    "看到": ["看到", "看见", "收到", "注意到"],
    "评论": ["评论", "留言", "消息", "询问"],
    "你好": ["你好", "您好", "嗨", "嘿"],
    "方法": ["方法", "办法", "方式", "方案"],
}


class TemplateVariation:
    """模板变体生成器：从模板池随机选取并可选同义词替换"""

    @staticmethod
    def pick_random_from_pool(templates: list[dict], trigger: str = None) -> str:
        """
        从模板池中随机选取一条。
        如果指定 trigger，优先匹配 trigger 相关模板；
        若多个匹配，随机选一个。
        """
        if not templates:
            return ""

        if trigger:
            # 收集匹配 trigger 的模板
            matched = [
                t for t in templates
                if str(t.get("trigger", "")).strip() == trigger
            ]
            if matched:
                chosen = random.choice(matched)
                return str(chosen.get("content", ""))

        # 无 trigger 匹配 → 从全部模板随机选
        chosen = random.choice(templates)
        return str(chosen.get("content", ""))

    @staticmethod
    def apply_synonym_variation(text: str, intensity: float = 0.3) -> str:
        """
        对文本进行轻度同义词替换，增加多样性。
        intensity: 0-1，替换概率（0.3 表示约 30% 可替换词会被替换）
        """
        words = list(SYNONYM_MAP.keys())
        random.shuffle(words)
        result = text
        for word in words:
            if random.random() < intensity:
                replacement = random.choice(SYNONYM_MAP[word])
                if replacement != word:
                    result = result.replace(word, replacement)
        return result

    @staticmethod
    def add_random_emoji(text: str, probability: float = 0.3) -> str:
        """随机追加表情符号"""
        if random.random() < probability:
            emojis = ["😊", "🙏", "❤️", "✨", "💪", "🌹", "😄", "👍"]
            text = text + random.choice(emojis)
        return text


class MaterialManager:
    def __init__(self, excel_path: str = None):
        path = Path(excel_path or cfg.MATERIALS_DIR / "materials.xlsx")
        if not path.exists():
            self._create_template(path)
        try:
            self.wb = openpyxl.load_workbook(path)
        except Exception:
            self._create_template(path)
            self.wb = openpyxl.load_workbook(path)
        self._excel_path = path
        self._daily_counters: dict[str, dict[int, int]] = {
            "copywriting": {},
            "images": {},
        }
        self._last_pick_day: str = ""

    def _create_template(self, path: Path):
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "评论文案"
        ws1.append(["id", "category", "content", "priority", "daily_limit", "enabled"])
        ws1.append([1, "效果展示", "用了这个方法，白斑真的淡了！前后对比太明显了", "high", 20, "✅"])
        ws1.append([2, "经验分享", "我也是白斑困扰多年，终于找到了方法", "high", 15, "✅"])
        ws1.append([3, "互动引导", "有同样困扰的姐妹看过来，这个方法真的有用", "medium", 10, "✅"])
        ws2 = wb.create_sheet("对比图片")
        ws2.append(["id", "name", "before_path", "after_path", "category", "daily_limit", "enabled"])
        ws2.append([1, "手部对比", "images/before_hand.jpg", "images/after_hand.jpg", "手部", 15, "✅"])
        ws2.append([2, "面部对比", "images/before_face.jpg", "images/after_face.jpg", "面部", 15, "✅"])
        ws2.append([3, "腿部对比", "images/before_leg.jpg", "images/after_leg.jpg", "腿部", 10, "✅"])
        ws3 = wb.create_sheet("私信模板")
        ws3.append(["id", "trigger", "content", "enabled"])
        ws3.append([1, "默认", "看到你评论问我，我是吃药加照光弄好的", "✅"])
        ws3.append([2, "追问", "具体就是口服药+定期照光，坚持了半年左右，你可以试试", "✅"])
        ws4 = wb.create_sheet("回复模板")
        ws4.append(["id", "trigger_keywords", "reply_type", "content", "enabled"])
        ws4.append([1, "怎么治,如何治,什么方法", "text", "吃药加照光", "✅"])
        ws4.append([2, "真的吗,有用吗,效果", "text", "真的！坚持就会有效果", "✅"])
        ws4.append([3, "*", "emoji", "😊👍", "✅"])
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)

    def _reset_daily_counters_if_new_day(self):
        today = time.strftime("%Y-%m-%d")
        if self._last_pick_day != today:
            self._daily_counters = {"copywriting": {}, "images": {}}
            self._last_pick_day = today

    def _read_sheet(self, sheet_name: str) -> list[dict]:
        if sheet_name not in self.wb.sheetnames:
            return []
        ws = self.wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return []
        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
        result = []
        for row in rows[1:]:
            item = dict(zip(headers, row))
            enabled = str(item.get("enabled", "✅")).strip()
            if enabled not in ("✅", "True", "true", "1", "是"):
                continue
            result.append(item)
        return result

    def get_copywritings(self) -> list[dict]:
        return self._read_sheet("评论文案")

    def get_image_pairs(self) -> list[dict]:
        return self._read_sheet("对比图片")

    def get_dm_templates(self) -> list[dict]:
        return self._read_sheet("私信模板")

    def get_reply_templates(self) -> list[dict]:
        return self._read_sheet("回复模板")

    def pick_copywriting(self) -> Optional[dict]:
        self._reset_daily_counters_if_new_day()
        items = self.get_copywritings()
        if not items:
            return None
        weights = {"high": 5, "medium": 3, "low": 1}
        available = [
            it for it in items
            if self._daily_counters["copywriting"].get(it["id"], 0)
               < int(it.get("daily_limit", 999))
        ]
        if not available:
            self._daily_counters["copywriting"] = {}
            available = items
        chosen = random.choices(
            available,
            weights=[weights.get(str(it.get("priority", "medium")), 3)
                     for it in available],
            k=1
        )[0]
        cid = chosen["id"]
        self._daily_counters["copywriting"][cid] = \
            self._daily_counters["copywriting"].get(cid, 0) + 1
        return chosen

    def pick_image_pair(self) -> Optional[dict]:
        self._reset_daily_counters_if_new_day()
        items = self.get_image_pairs()
        if not items:
            return None
        available = [
            it for it in items
            if self._daily_counters["images"].get(it["id"], 0)
               < int(it.get("daily_limit", 999))
        ]
        if not available:
            self._daily_counters["images"] = {}
            available = items
        chosen = random.choice(available)
        cid = chosen["id"]
        self._daily_counters["images"][cid] = \
            self._daily_counters["images"].get(cid, 0) + 1
        return chosen

    def pick_dm(self, trigger: str = "默认") -> str:
        """
        从私信模板池随机选取一条。
        优先匹配 trigger，多个匹配时随机选一个（避免重复）。
        可选应用轻度同义词替换增加多样性。
        """
        templates = self.get_dm_templates()
        if not templates:
            return "看你评论问我，我是吃药加光照弄好的"

        # 收集所有匹配的模板
        if trigger != "默认":
            matched = [
                t for t in templates
                if str(t.get("trigger", "")).strip() == trigger
            ]
        else:
            matched = templates

        if not matched:
            matched = templates

        # 随机选取
        chosen = random.choice(matched)
        text = str(chosen.get("content", ""))

        # 30% 概率应用轻度同义词替换
        if random.random() < 0.3:
            text = TemplateVariation.apply_synonym_variation(text, intensity=0.2)

        return text

    def pick_reply(self, user_comment: str) -> str:
        """
        根据用户评论内容匹配回复模板，随机选取一条。
        多个匹配时随机选（避免重复），无匹配返回默认表情包。
        """
        templates = self.get_reply_templates()
        if not templates:
            return "😊👍"

        # 收集所有关键词匹配的模板
        matched = []
        for t in templates:
            keywords = str(t.get("trigger_keywords", "")).split(",")
            if "*" in keywords:
                continue  # 通配符作为兜底
            for kw in keywords:
                if kw.strip() in user_comment:
                    matched.append(t)
                    break  # 一个模板只加一次

        # 有匹配 → 随机选一条
        if matched:
            chosen = random.choice(matched)
            text = str(chosen.get("content", ""))
            # 对文本类型应用轻度变体
            if str(chosen.get("reply_type", "")).strip() == "text":
                if random.random() < 0.25:
                    text = TemplateVariation.apply_synonym_variation(
                        text, intensity=0.2
                    )
                text = TemplateVariation.add_random_emoji(text, probability=0.3)
            return text

        # 无匹配 → 随机选通配符兜底
        fallback = [
            t for t in templates
            if "*" in str(t.get("trigger_keywords", ""))
        ]
        if fallback:
            chosen = random.choice(fallback)
            return str(chosen.get("content", ""))

        return "😊👍"

    def reload(self):
        self.wb = openpyxl.load_workbook(self._excel_path)
